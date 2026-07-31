import openpyxl

wb = openpyxl.load_workbook('/Users/djchoi81/Downloads/실손보험_세대별_자기부담계산기_5.xlsx', data_only=True)
ws = wb['세대별변천사_원본표']

print(f"Sheet: {ws.title} ({ws.dimensions})")
for r in range(1, ws.max_row + 1):
    row_vals = []
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=r, column=c)
        if cell.value is not None:
            row_vals.append(f"{cell.coordinate}:{cell.value}")
    if row_vals:
        print(" | ".join(row_vals))
