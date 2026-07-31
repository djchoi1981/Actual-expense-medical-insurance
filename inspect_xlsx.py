import openpyxl

wb = openpyxl.load_workbook('/Users/djchoi81/Downloads/실손보험_세대별_자기부담계산기_5.xlsx', data_only=False)
print("Sheets:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\nSheet: {name}, dimensions: {ws.dimensions}")
